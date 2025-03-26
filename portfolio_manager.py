# portfolio_manager.py

import pandas as pd
import requests
from bs4 import BeautifulSoup
import smtplib
import json
from email.mime.text import MIMEText
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

CONFIG_FILE = "config.json"
PORTFOLIO_FILE = "portfolio.csv"
EXCEL_FILE = "portfolio_output.xlsx"


def load_config():
    try:
        with open(CONFIG_FILE, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"❌ Configuration file '{CONFIG_FILE}' not found.")
        exit(1)
    except json.JSONDecodeError:
        print(f"❌ Configuration file '{CONFIG_FILE}' is not valid JSON.")
        exit(1)


def load_portfolio(file_path):
    try:
        df = pd.read_csv(file_path)
        required_cols = {'ticker', 'shares', 'target_pct'}
        if not required_cols.issubset(df.columns):
            print(f"❌ Portfolio file must contain columns: {required_cols}")
            exit(1)
        return df
    except FileNotFoundError:
        print(f"❌ Portfolio file '{file_path}' not found.")
        exit(1)
    except pd.errors.ParserError:
        print(f"❌ Portfolio file '{file_path}' is not a valid CSV.")
        exit(1)


def get_price_from_yahoo(ticker):
    url = f"https://finance.yahoo.com/quote/{ticker}"
    headers = {"User-Agent": "Mozilla/5.0"}
    r = requests.get(url, headers=headers)
    soup = BeautifulSoup(r.text, 'html.parser')
    price_span = soup.find("fin-streamer", {"data-field": "regularMarketPrice"})
    if price_span:
        return float(price_span.text.replace(',', ''))
    raise ValueError(f"Couldn't get price for {ticker}")


def fetch_prices(tickers):
    return [get_price_from_yahoo(ticker) for ticker in tickers]


def calculate_rebalance(df):
    df['price'] = fetch_prices(df['ticker'].tolist())
    df['value'] = df['shares'] * df['price']
    total_value = df['value'].sum()

    df['actual_pct'] = df['value'] / total_value
    df['deviation'] = df['actual_pct'] - df['target_pct']
    df['rebalance_flag'] = df['deviation'].abs() > 0.05

    df['target_value'] = df['target_pct'] * total_value
    df['rebalance_amount'] = df['target_value'] - df['value']
    df['shares_to_trade'] = (df['rebalance_amount'] / df['price']).round()

    return df, total_value


def save_to_excel(df, total_value):
    snapshot = df[['ticker', 'price', 'value']].copy()
    snapshot['date'] = datetime.today().strftime('%Y-%m-%d')
    snapshot['total_value'] = total_value

    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        df[['ticker', 'shares', 'price', 'value', 'actual_pct', 'target_pct',
            'deviation', 'rebalance_flag', 'shares_to_trade']].to_excel(writer, sheet_name='Rebalance Report', index=False)
        snapshot.to_excel(writer, sheet_name='History', index=False)

    print(f"📄 Portfolio data saved to {EXCEL_FILE}")
    add_chart_to_history(EXCEL_FILE)


def add_chart_to_history(file_path):
    wb = load_workbook(file_path)
    ws = wb['History']

    values_col = None
    dates_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "total_value":
            values_col = idx
        if cell.value == "date":
            dates_col = idx

    if not values_col or not dates_col:
        print("⚠️ Unable to find columns for chart.")
        return

    data = Reference(ws, min_col=values_col, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=dates_col, min_row=2, max_row=ws.max_row)

    chart = LineChart()
    chart.title = "Portfolio Value Over Time"
    chart.y_axis.title = "Total Value"
    chart.x_axis.title = "Date"

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    for obj in ws._charts:
        ws._charts.remove(obj)

    ws.add_chart(chart, "H2")
    wb.save(file_path)
    print(f"📊 Chart added to {file_path}")


def send_email(subject, body, config):
    try:
        msg = MIMEText(body)
        msg['Subject'] = subject
        msg['From'] = config["sender_email"]
        msg['To'] = config["recipient_email"]

        with smtplib.SMTP(config["smtp_server"], config["smtp_port"]) as server:
            server.starttls()
            server.login(config["sender_email"], config["email_password"])
            server.sendmail(config["sender_email"], config["recipient_email"], msg.as_string())
            print("📧 Email sent.")
    except Exception as e:
        print(f"❌ Failed to send email: {e}")


def check_and_notify(df, config):
    rebalance_needed = df[df['rebalance_flag']]
    if config.get("email_enabled") and not rebalance_needed.empty:
        body = rebalance_needed[['ticker', 'deviation', 'shares_to_trade']].to_string(index=False)
        send_email("⚖️ Portfolio Rebalance Alert", body, config)


def main():
    print(f"\n📊 Portfolio Rebalancing — {datetime.today().date()}")
    config = load_config()
    portfolio = load_portfolio(PORTFOLIO_FILE)
    df, total_value = calculate_rebalance(portfolio)

    print("\n🔍 Rebalance Check (5% Threshold Rule):\n")
    print(df[['ticker', 'shares', 'price', 'value', 'actual_pct', 'target_pct',
              'deviation', 'rebalance_flag', 'shares_to_trade']])

    save_to_excel(df, total_value)
    check_and_notify(df, config)


if __name__ == "__main__":
    main()